# save as myopia_diagnosis_complete_final.py
import pandas as pd
import numpy as np
import json
import torch
import random
import os
import traceback
import gc
import hashlib
import time
import warnings
import matplotlib.pyplot as plt
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
from collections import defaultdict
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.preprocessing import LabelEncoder, StandardScaler, label_binarize
from sklearn.metrics import (classification_report, confusion_matrix, roc_curve, auc, 
                           accuracy_score, precision_score, recall_score, f1_score)
import joblib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# 尝试导入 xgboost，如果失败则设为可选
try:
    from xgboost import XGBClassifier
    XGBOOST_AVAILABLE = True
    print("✅ XGBoost available")
except ImportError:
    XGBOOST_AVAILABLE = False
    print("⚠️ XGBoost not installed, skipping XGBoost model")

def set_random_seeds(seed=42):
    """设置所有随机种子保证可重复性"""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    print(f"✅ Random seeds set to: {seed}")

# 调用设置随机种子
set_random_seeds(42)

# 检查是否有GPU
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Device: {device}")

class PatientIDGenerator:
    """Patient ID Generator (generates 16-bit unique identifier)"""
    
    @staticmethod
    def generate_patient_id(patient_info: Dict[str, Any]) -> str:
        """
        Generate a 16-bit unique identifier similar to Docker container ID
        Format: 4 groups of 4 hex digits, e.g.: a1b2-c3d4-e5f6-g7h8
        """
        # Create unique string
        unique_str = f"{patient_info.get('姓名', '')}_{patient_info.get('性别', '')}_{patient_info.get('年龄', '')}_{time.time_ns()}"
        
        # Generate unique value using SHA256 hash
        hash_obj = hashlib.sha256(unique_str.encode())
        hex_digest = hash_obj.hexdigest()[:16]  # Take first 16 characters
        
        # Format into 4 groups of 4
        formatted_id = '-'.join([hex_digest[i:i+4] for i in range(0, 16, 4)])
        
        return formatted_id.upper()

class ModelPerformanceTracker:
    """Model Performance Tracker"""
    
    def __init__(self):
        self.history = []
        self.best_models = {}
        self.current_iteration = 0
        
    def record_performance(self, model_name: str, metrics: Dict, iteration: int):
        """Record model performance"""
        record = {
            'model_name': model_name,
            'iteration': iteration,
            'timestamp': datetime.now().isoformat(),
            **metrics
        }
        self.history.append(record)
        
        # Update best model
        if model_name not in self.best_models or metrics['accuracy'] > self.best_models[model_name]['metrics']['accuracy']:
            self.best_models[model_name] = {
                'iteration': iteration,
                'metrics': metrics,
                'timestamp': record['timestamp']
            }
            
    def get_best_models(self) -> Dict:
        """Get best model information"""
        return self.best_models
    
    def save_history(self, filepath: str = "model_history.json"):
        """Save history"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.history, f, ensure_ascii=False, indent=2)
            
    def load_history(self, filepath: str = "model_history.json"):
        """Load history"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                self.history = json.load(f)
        except FileNotFoundError:
            print(f"History file {filepath} not found")

class MyopiaRuleClassifier:
    """Rule-based Diagnosis System"""
    AXIAL_LENGTH_THRESHOLD = 26.00
    AL_CR_CONSTANT = 337.5

    def __init__(self, use_cycloplegic: bool = False):
        self.use_cycloplegic = use_cycloplegic
        self._load_medical_reference_data()

    def _load_medical_reference_data(self):
        self.axial_ref = pd.DataFrame({
            'age': [6,7,8,9,10,11,12,13,14,15],
            'min_len': [20.93,21.07,21.30,21.45,21.60,21.71,21.79,22.07,21.92,22.10],
            'max_len': [23.98,24.04,24.27,24.46,24.67,24.80,24.84,24.65,24.82,24.68],
            'mean': [22.46,22.68,22.90,23.05,23.22,23.38,23.52,23.62,23.72,23.39]
        })
        self.corneal_ref = pd.DataFrame({
            'age': [6,7,8,9,10,11,12,13,14,15],
            'min_curv': [7.93,7.09,7.42,7.41,7.41,7.42,7.39,7.39,7.36,7.40],
            'max_curv': [8.45,8.70,8.41,8.43,8.43,8.41,8.43,8.46,8.47,8.38]
        })
        self.vision_ref = {3:0.5,4:0.6,5:0.8,6:1.0}

    def _check_axial_length(self, age: int, al: float) -> str:
        ref = self.axial_ref[self.axial_ref['age'] == age]
        if ref.empty:
            return "Age out of reference range"
        min_len = ref['min_len'].values[0]
        max_len = ref['max_len'].values[0]
        if al < min_len:
            return f"Axial length too short (below {min_len:.2f}mm)"
        elif al > max_len:
            return f"Axial length too long (above {max_len:.2f}mm)"
        else:
            return f"Normal range ({min_len:.2f}-{max_len:.2f}mm)"

    def calculate_al_cr_ratio(self, al: float, corneal_curv: float) -> float:
        return round((al * self.jiaomoqulv(corneal_curv)) / self.AL_CR_CONSTANT, 2)

    def jiaomoqulv(self, curv: float) -> float:
        return self.AL_CR_CONSTANT / curv

    def _check_corneal_curvature(self, age: int, corneal_curv: float) -> str:
        ref = self.corneal_ref[self.corneal_ref['age'] == age]
        if ref.empty:
            return "Age out of reference range"
        min_curv = ref['min_curv'].values[0]
        max_curv = ref['max_curv'].values[0]
        if corneal_curv < min_curv:
            return f"Corneal curvature too steep (below {min_curv:.2f}D)"
        elif corneal_curv > max_curv:
            return f"Corneal curvature too flat (above {max_curv:.2f}D)"
        else:
            return f"Normal range ({min_curv:.2f}-{max_curv:.2f}D)"

    def _check_vision(self, age: int, va: float) -> str:
        if age in self.vision_ref:
            standard_va = self.vision_ref[age]
            if va < standard_va:
                return f"Vision development lag (age {age} standard: {standard_va}, current: {va})"
            else:
                return f"Normal vision (meets age {age} standard)"
        else:
            return "Age out of vision reference range"

    def diagnose(self, patient_data: Dict[str, float]) -> Dict:
        required_fields = ['age','se','al','corneal_curv','va','is_cycloplegic']
        if any(field not in patient_data for field in required_fields):
            missing = [f for f in required_fields if f not in patient_data]
            raise ValueError(f"Missing required fields: {missing}")
        
        diagnosis = {
            'axial_check': self._check_axial_length(patient_data['age'], patient_data['al']),
            'corneal_curv_check': self._check_corneal_curvature(patient_data['age'], patient_data['corneal_curv']),
            'vision_check': self._check_vision(patient_data['age'], patient_data['va']),
            'al_cr_ratio': self.calculate_al_cr_ratio(patient_data['al'], patient_data['corneal_curv']),
        }
        
        se = patient_data['se']
        age = patient_data['age']
        al = patient_data['al']
        corneal_curv = patient_data['corneal_curv']
        
        # Get reference values for age
        axial_ref = self.axial_ref[self.axial_ref['age'] == age]
        corneal_ref = self.corneal_ref[self.corneal_ref['age'] == age]
        
        # Calculate deviation indicators
        if not axial_ref.empty:
            al_deviation = (al - axial_ref['mean'].values[0])/axial_ref['mean'].values[0] * 100
        else:
            al_deviation = 0
            
        if not corneal_ref.empty:
            corneal_mean = (corneal_ref['min_curv'].values[0] + corneal_ref['max_curv'].values[0]) / 2
            corneal_deviation = corneal_curv - corneal_mean
        else:
            corneal_deviation = 0
        
        # Adjusted diagnosis logic - return English diagnoses directly
        se_threshold = -0.50 if patient_data['is_cycloplegic'] else -0.75
        if se >= 0.75:
            diagnosis['stage'] = "Hyperopia"  # 远视
        elif 0.75 > se >= se_threshold:
            # Emmetropia needs to consider axial length/corneal curvature deviation
            if abs(al_deviation) < 5 and abs(corneal_deviation) < 0.2:
                diagnosis['stage'] = "Emmetropia"  # 正视
            else:
                diagnosis['stage'] = "Pre-myopia"  # 近视前期（修改：疑似近视前期 -> 近视前期）
        elif se_threshold > se >= -3.25:
            diagnosis['stage'] = "Mild Myopia" if al_deviation < 10 else "Mild Myopia (Long Axial)"
        elif -3.25 > se >= -6.00:
            diagnosis['stage'] = "Moderate Myopia"
        elif se < -6.00 or (al >= self.AXIAL_LENGTH_THRESHOLD and al_deviation >= 15):
            diagnosis['stage'] = "High Myopia"
            diagnosis['warning'] = "Need to be alert to pathological myopia risk"
        else:
            diagnosis['stage'] = "Unclassified"
        
        return diagnosis

class AutoRepairingModel:
    """Auto-repairing model (automatically adjusts when performance declines)"""
    
    def __init__(self, base_model, model_name: str):
        self.base_model = base_model
        self.model_name = model_name
        self.performance_history = []
        self.fail_count = 0
        self.repair_count = 0
        
    def train_with_repair(self, X_train, y_train, X_val, y_val, max_retries=3):
        """Training with repair functionality"""
        for attempt in range(max_retries):
            try:
                # Train model
                self.base_model.fit(X_train, y_train)
                
                # Validate performance
                y_pred = self.base_model.predict(X_val)
                accuracy = accuracy_score(y_val, y_pred)
                
                # Check for overfitting
                if len(self.performance_history) >= 2:
                    last_accuracy = self.performance_history[-1]['accuracy']
                    if accuracy < last_accuracy * 0.8:  # Performance drop > 20%
                        print(f"  ⚠️ {self.model_name} performance drop ({accuracy:.2%} < {last_accuracy:.2%}), attempting repair...")
                        self._apply_repair()
                        continue
                
                # Record performance
                self.performance_history.append({
                    'attempt': attempt,
                    'accuracy': accuracy,
                    'timestamp': datetime.now().isoformat()
                })
                
                return accuracy
                
            except Exception as e:
                self.fail_count += 1
                print(f"  ❌ {self.model_name} training failed (attempt {attempt+1}/{max_retries}): {e}")
                
                if attempt < max_retries - 1:
                    self._apply_repair()
                else:
                    raise
        
        return 0.0
    
    def _apply_repair(self):
        """Apply repair strategy"""
        self.repair_count += 1
        
        if self.model_name == 'RandomForest':
            # RandomForest repair: increase tree depth or number
            if hasattr(self.base_model, 'n_estimators'):
                self.base_model.n_estimators = min(self.base_model.n_estimators * 2, 200)
            if hasattr(self.base_model, 'max_depth'):
                self.base_model.max_depth = self.base_model.max_depth + 2 if self.base_model.max_depth else 10
                
        elif self.model_name == 'XGBoost' and XGBOOST_AVAILABLE:
            # XGBoost repair: adjust learning rate
            if hasattr(self.base_model, 'learning_rate'):
                self.base_model.learning_rate = max(self.base_model.learning_rate * 0.8, 0.01)
        
        elif self.model_name == 'LogisticRegression':
            # LogisticRegression repair: increase iterations
            if hasattr(self.base_model, 'max_iter'):
                self.base_model.max_iter = min(self.base_model.max_iter * 2, 5000)
        
        elif self.model_name == 'SVM':
            # SVM repair: adjust C parameter
            if hasattr(self.base_model, 'C'):
                self.base_model.C = self.base_model.C * 1.5
                
        elif self.model_name == 'GradientBoosting':
            # GradientBoosting repair: adjust learning rate
            if hasattr(self.base_model, 'learning_rate'):
                self.base_model.learning_rate = max(self.base_model.learning_rate * 0.8, 0.01)
                
        print(f"  🔧 {self.model_name} applied repair strategy #{self.repair_count}")

class MultiModelTraining:
    """Multi-model training and comparison"""
    
    def __init__(self):
        self.models = {}
        self.scaler = StandardScaler()
        self.feature_columns = None
        self.performance_tracker = ModelPerformanceTracker()
        
    def prepare_features(self, data_df):
        """Prepare features"""
        print("  Preparing features...")
        
        # Basic features
        age_se_mean = data_df.groupby('年龄')['验光'].mean().to_dict()
        features = []
        
        for _, row in data_df.iterrows():
            # Replace direct refraction value with difference from age group mean
            se_diff = row['验光'] - age_se_mean.get(row['年龄'], row['验光'])
            feature_vector = [
                row['年龄'],
                1 if row['性别'] == '女' else 0,
                row['视力'],
                se_diff,
                row['眼轴'],
                row['轴率比'],
                row['角膜曲率'],
                row['眼轴'] * (337.5 / row['角膜曲率']) / 337.5,
                abs(se_diff)
            ]
            features.append(feature_vector)
        
        # Update feature column names
        self.feature_columns = ['Age', 'Gender_Female', 'Vision', 'Refraction_AgeDiff', 'Axial_Length', 
                              'Axial_Ratio', 'Corneal_Curvature', 'AL_CR_Ratio', 'Refraction_Diff_Abs']
        
        features_array = np.array(features)
        
        # Standardize features
        if len(features_array) > 1:
            features_array = self.scaler.fit_transform(features_array)
        
        return features_array
    
    def initialize_models(self):
        """Initialize all models"""
        print("  Initializing machine learning models...")
        
        # 1. Random Forest
        self.models['RandomForest'] = AutoRepairingModel(
            RandomForestClassifier(
                n_estimators=100,
                max_depth=10,
                random_state=42,
                class_weight='balanced'
            ),
            'RandomForest'
        )
        
        # 2. Logistic Regression
        self.models['LogisticRegression'] = AutoRepairingModel(
            LogisticRegression(
                max_iter=1000,
                random_state=42,
                multi_class='ovr',
                class_weight='balanced'
            ),
            'LogisticRegression'
        )
        
        # 3. Support Vector Machine
        self.models['SVM'] = AutoRepairingModel(
            SVC(
                kernel='rbf',
                probability=True,
                random_state=42,
                class_weight='balanced'
            ),
            'SVM'
        )
        
        # 4. Gradient Boosting
        self.models['GradientBoosting'] = AutoRepairingModel(
            GradientBoostingClassifier(
                n_estimators=100,
                learning_rate=0.1,
                max_depth=5,
                random_state=42
            ),
            'GradientBoosting'
        )
        
        # 5. XGBoost (optional)
        if XGBOOST_AVAILABLE:
            self.models['XGBoost'] = AutoRepairingModel(
                XGBClassifier(
                    n_estimators=100,
                    learning_rate=0.1,
                    max_depth=5,
                    random_state=42,
                    use_label_encoder=False,
                    eval_metric='mlogloss'
                ),
                'XGBoost'
            )
        else:
            print("  ⚠️ Skipping XGBoost model (not installed)")
        
        print(f"  Initialized {len(self.models)} models: {list(self.models.keys())}")
    
    def train_and_evaluate_all(self, train_data, val_data, test_data, iteration=1):
        """Train and evaluate all models"""
        print(f"\n  --- Training iteration {iteration} ---")
        #for _ in range():
        
        # Prepare data
        X_train = self.prepare_features(train_data)
        y_train = train_data['诊断编码'].values
        
        X_val = self.prepare_features(val_data)
        y_val = val_data['诊断编码'].values
        
        X_test = self.prepare_features(test_data)
        y_test = test_data['诊断编码'].values
        
        results = {}
        
        for model_name, model_wrapper in self.models.items():
            print(f"\n  Training {model_name}...")
            
            try:
                # Train with repair functionality
                start_time = time.time()
                val_accuracy = model_wrapper.train_with_repair(X_train, y_train, X_val, y_val)
                training_time = time.time() - start_time
                
                # Test performance
                model = model_wrapper.base_model
                y_pred = model.predict(X_test)
                
                # Check if predict_proba method exists
                has_proba = hasattr(model, 'predict_proba')
                y_proba = model.predict_proba(X_test) if has_proba else None
                
                # Calculate various metrics
                metrics = {
                    'accuracy': accuracy_score(y_test, y_pred),
                    'precision': precision_score(y_test, y_pred, average='weighted', zero_division=0),
                    'recall': recall_score(y_test, y_pred, average='weighted', zero_division=0),
                    'f1_score': f1_score(y_test, y_pred, average='weighted', zero_division=0),
                    'val_accuracy': val_accuracy,
                    'training_time': training_time,
                    'repair_count': model_wrapper.repair_count,
                    'fail_count': model_wrapper.fail_count,
                    'has_proba': has_proba
                }
                
                # Record performance
                self.performance_tracker.record_performance(model_name, metrics, iteration)
                
                results[model_name] = {
                    'model': model,
                    'metrics': metrics,
                    'predictions': y_pred,
                    'probabilities': y_proba,
                    'y_test': y_test,
                    'X_test': X_test
                }
                
                print(f"    Accuracy: {metrics['accuracy']:.2%}")
                print(f"    F1 Score: {metrics['f1_score']:.2%}")
                print(f"    Training time: {training_time:.2f}s")
                print(f"    Repair count: {model_wrapper.repair_count}")
                
            except Exception as e:
                print(f"    ❌ {model_name} training failed: {e}")
                results[model_name] = {'error': str(e)}
        
        return results
    
    def compare_models(self, results):
        """Compare model performance"""
        print("\n  --- Model Performance Comparison ---")
        
        performance_table = []
        for model_name, result in results.items():
            if 'metrics' in result:
                metrics = result['metrics']
                performance_table.append({
                    'Model': model_name,
                    'Accuracy': f"{metrics['accuracy']:.2%}",
                    'F1-Score': f"{metrics['f1_score']:.2%}",
                    'Training Time': f"{metrics['training_time']:.2f}s",
                    'Repairs': metrics['repair_count']
                })
        
        # Create comparison table
        comparison_df = pd.DataFrame(performance_table)
        if not comparison_df.empty:
            print(comparison_df.to_string(index=False))
            
            # Find best model
            best_model = max(performance_table, key=lambda x: float(x['Accuracy'].rstrip('%')))
            print(f"\n  🏆 Best model: {best_model['Model']} (Accuracy: {best_model['Accuracy']})")
        else:
            print("    No comparable model results")
        
        return comparison_df
    
    def ensemble_predict(self, results, X):
        """Ensemble prediction (using average probabilities of all models)"""
        print("\n  Executing ensemble prediction...")
        
        probabilities_list = []
        model_count = 0
        
        for model_name, result in results.items():
            if 'probabilities' in result and result['probabilities'] is not None:
                probabilities_list.append(result['probabilities'])
                model_count += 1
        
        if probabilities_list:
            # Average probabilities of all models
            avg_probabilities = np.mean(probabilities_list, axis=0)
            ensemble_predictions = np.argmax(avg_probabilities, axis=1)
            
            print(f"    Used {model_count} models for ensemble")
            return ensemble_predictions, avg_probabilities
        else:
            print("    ⚠️ No models support probability prediction, cannot perform ensemble")
            return None, None
    
    def analyze_roc_for_all_models(self, results, label_encoder, output_dir="roc_analysis"):
        """Analyze ROC curves and AUC values for all models"""
        print(f"\n  --- Analyzing ROC curves and AUC for all models ---")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        roc_results = {}
        
        for model_name, result in results.items():
            if 'probabilities' not in result or result['probabilities'] is None:
                print(f"    ⚠️ {model_name} does not support probability prediction, skipping ROC analysis")
                continue
            
            if 'y_test' not in result:
                print(f"    ⚠️ {model_name} has no test data, skipping ROC analysis")
                continue
            
            try:
                print(f"    📊 Analyzing ROC curve for {model_name}...")
                
                # Get test data and probabilities
                y_test = result['y_test']
                y_proba = result['probabilities']
                
                # Get number of classes
                n_classes = y_proba.shape[1]
                
                # Get class names
                if label_encoder is not None and hasattr(label_encoder, 'classes_'):
                    # Get original diagnosis names
                    original_class_names = label_encoder.classes_
                    
                    # Create Chinese-English mapping dictionary
                    diagnosis_translation = {
                        # English to English (they should already be in English from the rule classifier)
                        "Hyperopia": "Hyperopia",
                        "Emmetropia": "Emmetropia", 
                        "Pre-myopia": "Pre-myopia",  # 修改：近视前期
                        "Mild Myopia": "Mild Myopia",
                        "Mild Myopia (Long Axial)": "Mild Myopia (Long Axial)",
                        "Moderate Myopia": "Moderate Myopia",
                        "High Myopia": "High Myopia",
                        "Unclassified": "Unclassified",
                        "Diagnosis Failed": "Diagnosis Failed"
                    }
                    
                    # Translate Chinese class names to English
                    class_names = []
                    for name in original_class_names:
                        # Since our rule classifier now returns English directly,
                        # we can just use the original names
                        if isinstance(name, str):
                            class_names.append(name)
                        else:
                            class_names.append(str(name))
                else:
                    # If no label encoder, use default Class numbering
                    class_names = [f'Class {i}' for i in range(n_classes)]
                
                # Calculate ROC curve and AUC for each class
                fpr = {}
                tpr = {}
                roc_auc = {}
                
                # Binarize labels for multi-class ROC
                y_test_bin = label_binarize(y_test, classes=range(n_classes))
                
                for i in range(n_classes):
                    fpr[i], tpr[i], _ = roc_curve(y_test_bin[:, i], y_proba[:, i])
                    roc_auc[i] = auc(fpr[i], tpr[i])
                
                # Calculate micro-average ROC
                fpr["micro"], tpr["micro"], _ = roc_curve(y_test_bin.ravel(), y_proba.ravel())
                roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
                
                # Calculate macro-average ROC
                all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))
                mean_tpr = np.zeros_like(all_fpr)
                for i in range(n_classes):
                    mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])
                mean_tpr /= n_classes
                fpr["macro"] = all_fpr
                tpr["macro"] = mean_tpr
                roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])
                
                # Plot ROC curve
                plt.figure(figsize=(12, 8))
                colors = plt.cm.rainbow(np.linspace(0, 1, n_classes))
                
                # Plot ROC curve for each class
                for i, color in zip(range(n_classes), colors):
                    if i < len(class_names):
                        label_name = class_names[i]
                    else:
                        label_name = f'Class {i}'
                    
                    plt.plot(fpr[i], tpr[i], color=color, lw=2,
                            label=f'{label_name} (AUC = {roc_auc[i]:.3f})')
                
                # Plot average ROC curves
                plt.plot(fpr["micro"], tpr["micro"],
                        label=f'Micro-average (AUC = {roc_auc["micro"]:.3f})',
                        color='deeppink', linestyle=':', linewidth=4)
                
                plt.plot(fpr["macro"], tpr["macro"],
                        label=f'Macro-average (AUC = {roc_auc["macro"]:.3f})',
                        color='navy', linestyle=':', linewidth=4)
                
                plt.plot([0, 1], [0, 1], 'k--', lw=2, label='Random (AUC = 0.5)')
                
                plt.xlim([0.0, 1.0])
                plt.ylim([0.0, 1.05])
                plt.xlabel('False Positive Rate (FPR)', fontsize=12)
                plt.ylabel('True Positive Rate (TPR)', fontsize=12)
                plt.title(f'{model_name} - Multi-class ROC Curve', fontsize=14, fontweight='bold')
                plt.legend(loc="lower right", fontsize=9)
                plt.grid(True, alpha=0.3)
                
                # Save image
                roc_file = os.path.join(output_dir, f"{model_name}_ROC_Curve.png")
                plt.tight_layout()
                plt.savefig(roc_file, dpi=300, bbox_inches='tight')
                plt.close()
                
                print(f"      ROC curve saved to: {roc_file}")
                
                # Save AUC results
                roc_results[model_name] = {
                    'class_auc': roc_auc,
                    'class_names_english': class_names,
                    'micro_auc': roc_auc['micro'],
                    'macro_auc': roc_auc['macro']
                }
                
                # Print AUC summary
                print(f"      AUC Summary:")
                print(f"        Micro-average AUC: {roc_auc['micro']:.3f}")
                print(f"        Macro-average AUC: {roc_auc['macro']:.3f}")
                for i in range(n_classes):
                    if i < len(class_names):
                        label_name = class_names[i]
                    else:
                        label_name = f'Class {i}'
                    print(f"        {label_name} AUC: {roc_auc[i]:.3f}")
                    
            except Exception as e:
                print(f"    ❌ ROC analysis failed for {model_name}: {e}")
                roc_results[model_name] = {'error': str(e)}
        
        # Save all ROC results
        roc_summary_file = os.path.join(output_dir, "ROC_AUC_Summary.json")
        with open(roc_summary_file, 'w', encoding='utf-8') as f:
            # Convert numpy types to Python types
            serializable_results = {}
            for model_name, result in roc_results.items():
                if 'error' in result:
                    serializable_results[model_name] = result
                else:
                    serializable_results[model_name] = {
                        'micro_auc': float(result['micro_auc']),
                        'macro_auc': float(result['macro_auc']),
                        'class_names_english': result.get('class_names_english', []),
                        'class_auc': {str(k): float(v) for k, v in result['class_auc'].items() if k not in ['micro', 'macro']}
                    }
            
            json.dump(serializable_results, f, ensure_ascii=False, indent=2)
        
        print(f"\n  ✅ ROC analysis completed, results saved to: {output_dir}")
        return roc_results
    
    def save_all_models(self, base_dir="saved_models"):
        """Save all models"""
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        
        for model_name, model_wrapper in self.models.items():
            try:
                model_path = os.path.join(base_dir, f"{model_name}_model.joblib")
                joblib.dump(model_wrapper.base_model, model_path)
                print(f"  Saved {model_name} model")
            except Exception as e:
                print(f"  Failed to save {model_name}: {e}")
        
        # Save feature processor
        scaler_path = os.path.join(base_dir, "scaler.joblib")
        joblib.dump(self.scaler, scaler_path)
        
        # Save performance history
        self.performance_tracker.save_history(os.path.join(base_dir, "model_history.json"))
        
        # Save label encoder (if available)
        label_encoder_path = os.path.join(base_dir, "label_encoder.joblib")
        if hasattr(self, 'label_encoder') and self.label_encoder is not None:
            joblib.dump(self.label_encoder, label_encoder_path)
            print(f"  Saved label encoder")
        
        print(f"\n  All models saved to: {base_dir}")

class EnhancedDataProcessor:
    """Enhanced Data Processor"""
    
    def __init__(self):
        self.rule_classifier = MyopiaRuleClassifier()
        self.label_encoder = LabelEncoder()
        self.patient_id_map = {}  # Mapping: original info -> patient ID
        
    def load_data(self, excel_path: str):
        print("1. Loading data...")
        df = pd.read_excel(excel_path, header=None)  # Read without headers
        print(f"  Raw data: {len(df)} records")
        print(f"  Number of columns: {len(df.columns)}")
        return df
    
    def process_data(self, df):
        """Process data"""
        print("2. Processing data...")
        
        processed_data = []
        
        for idx, row in df.iterrows():
            # Check if necessary columns exist
            if len(row) < 11:
                print(f"  Row {idx+1} incomplete, skipping")
                continue
            
            # Process right eye
            if not pd.isna(row[3]) and not pd.isna(row[5]) and not pd.isna(row[7]) and not pd.isna(row[9]):
                try:
                    eye_data = self._process_single_eye(row, 'right')
                    processed_data.append(eye_data)
                except Exception as e:
                    print(f"  Failed to process right eye data (row {idx+1}): {e}")
            
            # Process left eye
            if not pd.isna(row[4]) and not pd.isna(row[6]) and not pd.isna(row[8]) and not pd.isna(row[10]):
                try:
                    eye_data = self._process_single_eye(row, 'left')
                    processed_data.append(eye_data)
                except Exception as e:
                    print(f"  Failed to process left eye data (row {idx+1}): {e}")
        
        data_df = pd.DataFrame(processed_data)
        print(f"  After processing: {len(data_df)} eye data points")
        return data_df
    
    def _process_single_eye(self, row, eye_side: str) -> Dict:
        """Process single eye data"""
        # Get data based on Excel column positions
        if eye_side == 'right':
            va = float(row[3])
            se_str = str(row[5])
            al = float(row[7])
            axial_ratio = float(row[9])
        else:
            va = float(row[4])
            se_str = str(row[6])
            al = float(row[8])
            axial_ratio = float(row[10])
        
        age_str = str(row[2])
        if '/' in age_str:
            years, _ = age_str.split('/')
            age = int(years)
        else:
            age = int(age_str)
        
        # Process refraction string
        se_str_clean = str(se_str).strip()
        if se_str_clean.startswith('+'):
            se = float(se_str_clean[1:])
        elif se_str_clean.startswith('-'):
            se = float(se_str_clean)
        else:
            se = float(se_str_clean)
        
        # Calculate corneal curvature
        corneal_curv = 45 - (axial_ratio - 3.0) * 10
        if corneal_curv < 38:
            corneal_curv = 38
        elif corneal_curv > 48:
            corneal_curv = 48
        
        # Create patient key
        gender_num = row[1] if len(row) > 1 else 0
        patient_key = f"{row[0]}_{gender_num}_{age_str}"
        
        return {
            '姓名': str(row[0]),
            '性别': '女' if gender_num == 1 else '男',
            '年龄': age,
            '眼别': '右眼' if eye_side == 'right' else '左眼',
            '视力': va,
            '验光': se,
            '眼轴': al,
            '轴率比': axial_ratio,
            '角膜曲率': round(corneal_curv, 2),
            '原始_年龄': age_str,
            '原始_验光': se_str_clean,
            'patient_key': patient_key
        }
    
    def generate_patient_ids(self, data_df):
        """Generate unique IDs for each patient"""
        print("3. Generating unique patient IDs...")
        
        unique_patients = data_df['patient_key'].unique()
        
        for patient_key in unique_patients:
            parts = patient_key.split('_')
            if len(parts) >= 3:
                patient_info = {
                    '姓名': parts[0],
                    '性别': parts[1],
                    '年龄': parts[2]
                }
                
                patient_id = PatientIDGenerator.generate_patient_id(patient_info)
                self.patient_id_map[patient_key] = patient_id
        
        # Add patient IDs to dataframe
        data_df['patient_id'] = data_df['patient_key'].map(self.patient_id_map)
        
        print(f"  Generated unique IDs for {len(unique_patients)} patients")
        print(f"  Example IDs: {list(self.patient_id_map.values())[:3]}")
        
        return data_df
    
    def run_rule_diagnosis(self, data_df):
        """Run rule-based diagnosis"""
        print("4. Generating rule-based diagnosis labels...")
        
        diagnoses = []
        for _, row in data_df.iterrows():
            try:
                patient_data = {
                    'age': row['年龄'],
                    'se': row['验光'],
                    'al': row['眼轴'],
                    'corneal_curv': row['角膜曲率'],
                    'va': row['视力'],
                    'is_cycloplegic': False
                }
                diagnosis = self.rule_classifier.diagnose(patient_data)
                diagnoses.append(diagnosis['stage'])
            except Exception as e:
                print(f"  Diagnosis failed (patient: {row['姓名']}, eye: {row['眼别']}): {e}")
                diagnoses.append("Diagnosis Failed")
        
        data_df['诊断结果'] = diagnoses
        
        # Encode diagnosis results
        valid_diagnoses = [d for d in diagnoses if d != "Diagnosis Failed"]
        if valid_diagnoses:
            self.label_encoder.fit(valid_diagnoses)
            data_df['诊断编码'] = data_df['诊断结果'].apply(
                lambda x: self.label_encoder.transform([x])[0] if x in self.label_encoder.classes_ else -1
            )
        else:
            data_df['诊断编码'] = -1
        
        print(f"  Diagnosis category distribution:")
        for label, count in data_df['诊断结果'].value_counts().items():
            print(f"    {label}: {count} ({count/len(data_df):.1%})")
        
        return data_df
    
    # 修改 EnhancedDataProcessor 类中的 split_data_by_patient_id 方法：

def split_data_by_patient_id(self, data_df, train_size=0.7, val_size=0.15, test_size=0.15, random_state=None):
    """按患者ID分割数据（添加随机种子参数）"""
    print("5. 按患者ID分割数据...")
    
    # 过滤掉诊断失败的数据
    valid_data = data_df[data_df['诊断编码'] != -1].copy()
    if len(valid_data) == 0:
        raise ValueError("没有有效的诊断数据可用于分割")
    
    unique_patients = valid_data['patient_id'].unique()
    
    # 分割患者ID - 使用传入的随机种子
    train_patients, temp_patients = train_test_split(
        unique_patients,
        test_size=(val_size + test_size),
        random_state=random_state  # 使用传入的随机种子
    )
    
    val_relative_size = val_size / (val_size + test_size)
    val_patients, test_patients = train_test_split(
        temp_patients,
        test_size=(1 - val_relative_size),
        random_state=random_state  # 使用相同的随机种子确保一致性
    )
    
    # 根据患者ID分配数据
    train_data = valid_data[valid_data['patient_id'].isin(train_patients)].copy()
    val_data = valid_data[valid_data['patient_id'].isin(val_patients)].copy()
    test_data = valid_data[valid_data['patient_id'].isin(test_patients)].copy()
    
    print(f"  训练集: {len(train_data)} 样本 ({len(train_data)/len(valid_data):.1%})")
    print(f"  验证集: {len(val_data)} 样本 ({len(val_data)/len(valid_data):.1%})")
    print(f"  测试集: {len(test_data)} 样本 ({len(test_data)/len(valid_data):.1%})")
    
    # 检查类别分布以确保多样性
    self._check_class_distribution(train_data, "训练集")
    self._check_class_distribution(val_data, "验证集")
    self._check_class_distribution(test_data, "测试集")
    
    return train_data, val_data, test_data

def _check_class_distribution(self, data, dataset_name):
    """检查数据集的类别分布"""
    if len(data) > 0:
        print(f"    {dataset_name}类别分布:")
        total = len(data)
        for diagnosis in data['诊断结果'].value_counts().index:
            count = (data['诊断结果'] == diagnosis).sum()
            percentage = count / total * 100
            print(f"      {diagnosis}: {count} ({percentage:.1f}%)")
    
    def save_datasets(self, train_data, val_data, test_data):
        """Save datasets"""
        print("6. Saving datasets...")
        
        train_data.to_excel("训练集.xlsx", index=False)
        val_data.to_excel("验证集.xlsx", index=False)
        test_data.to_excel("测试集.xlsx", index=False)
        
        print("  Datasets saved as Excel files")

class ExcelReportGenerator:
    """Excel Report Generator with Explanations"""
    
    @staticmethod
    def generate_training_report(history_file="model_history.json", 
                               iteration_file="iteration_results.json",
                               output_file="训练报告.xlsx"):
        """Generate Excel training report with explanations"""
        print(f"\n  Generating Excel training report...")
        
        try:
            # Create Excel workbook
            wb = Workbook()
            
            # 1. Training History Sheet with Explanation
            if os.path.exists(history_file):
                with open(history_file, 'r', encoding='utf-8') as f:
                    history_data = json.load(f)
                
                ws1 = wb.active
                ws1.title = "训练历史"
                
                # Add explanation at the top
                ws1.merge_cells('A1:I2')
                explanation_cell = ws1.cell(row=1, column=1, 
                    value="📊 训练历史表说明：\n"
                    "此表格记录了每个模型在每次训练迭代中的详细性能指标。\n"
                    "• 模型名称：使用的机器学习算法\n"
                    "• 迭代次数：训练轮次\n"
                    "• 时间戳：训练完成时间\n"
                    "• 准确率：模型在测试集上的预测准确率\n"
                    "• F1分数：综合考虑精确率和召回率的指标\n"
                    "• 精确率：正确预测的正例占所有预测为正例的比例\n"
                    "• 召回率：正确预测的正例占所有实际正例的比例\n"
                    "• 训练时间(秒)：模型训练耗时\n"
                    "• 修复次数：自修复功能触发的次数")
                explanation_cell.alignment = Alignment(wrap_text=True, vertical='center')
                explanation_cell.font = Font(bold=True, size=11)
                explanation_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                
                # Write headers (starting from row 4)
                headers = ["模型名称", "迭代次数", "时间戳", "准确率", "F1分数", 
                          "精确率", "召回率", "训练时间(秒)", "修复次数"]
                for col, header in enumerate(headers, 1):
                    cell = ws1.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data (starting from row 5)
                row = 5
                for record in history_data:
                    ws1.cell(row=row, column=1, value=record.get('model_name', ''))
                    ws1.cell(row=row, column=2, value=record.get('iteration', 0))
                    ws1.cell(row=row, column=3, value=record.get('timestamp', ''))
                    ws1.cell(row=row, column=4, value=record.get('accuracy', 0))
                    ws1.cell(row=row, column=5, value=record.get('f1_score', 0))
                    ws1.cell(row=row, column=6, value=record.get('precision', 0))
                    ws1.cell(row=row, column=7, value=record.get('recall', 0))
                    ws1.cell(row=row, column=8, value=record.get('training_time', 0))
                    ws1.cell(row=row, column=9, value=record.get('repair_count', 0))
                    row += 1
                
                # Adjust column widths
                for column in ws1.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws1.column_dimensions[column_letter].width = adjusted_width
                
                # Format numbers
                for row in ws1.iter_rows(min_row=5, max_row=row-1, min_col=4, max_col=8):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%' if cell.column <= 7 else '0.00'
            else:
                print(f"    ⚠️ Training history file {history_file} not found")
            
            # 2. Iteration Results Sheet with Explanation
            if os.path.exists(iteration_file):
                with open(iteration_file, 'r', encoding='utf-8') as f:
                    iteration_data = json.load(f)
                
                ws2 = wb.create_sheet("迭代结果")
                
                # Add explanation at the top
                ws2.merge_cells('A1:E2')
                explanation_cell = ws2.cell(row=1, column=1, 
                    value="📈 迭代结果表说明：\n"
                    "此表格汇总了每次训练迭代中所有模型的性能表现。\n"
                    "• 迭代次数：训练轮次编号\n"
                    "• 模型名称：使用的机器学习算法\n"
                    "• 准确率：模型在测试集上的预测准确率\n"
                    "• F1分数：综合考虑精确率和召回率的指标\n"
                    "• 训练时间(秒)：模型训练耗时\n\n"
                    "💡 观察方法：\n"
                    "1. 查看模型在多次迭代中的稳定性\n"
                    "2. 比较不同模型在同一迭代中的表现\n"
                    "3. 观察自修复功能是否提高了性能")
                explanation_cell.alignment = Alignment(wrap_text=True, vertical='center')
                explanation_cell.font = Font(bold=True, size=11)
                explanation_cell.fill = PatternFill(start_color="E6FAE6", end_color="E6FAE6", fill_type="solid")
                
                # Write headers (starting from row 4)
                headers = ["迭代次数", "模型名称", "准确率", "F1分数", "训练时间(秒)"]
                for col, header in enumerate(headers, 1):
                    cell = ws2.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data (starting from row 5)
                row = 5
                for iteration in iteration_data:
                    iteration_num = iteration.get('iteration', 0)
                    model_performance = iteration.get('model_performance', {})
                    
                    for model_name, metrics in model_performance.items():
                        ws2.cell(row=row, column=1, value=iteration_num)
                        ws2.cell(row=row, column=2, value=model_name)
                        ws2.cell(row=row, column=3, value=metrics.get('accuracy', 0))
                        ws2.cell(row=row, column=4, value=metrics.get('f1_score', 0))
                        ws2.cell(row=row, column=5, value=metrics.get('training_time', 0))
                        row += 1
                
                # Adjust column widths
                for column in ws2.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws2.column_dimensions[column_letter].width = adjusted_width
                
                # Format numbers
                for row in ws2.iter_rows(min_row=5, max_row=row-1, min_col=3, max_col=5):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%' if cell.column <= 4 else '0.00'
            else:
                print(f"    ⚠️ Iteration results file {iteration_file} not found")
            
            # 3. Model Comparison Sheet with Explanation
            ws3 = wb.create_sheet("模型对比")
            
            # Analyze best models
            best_models = {}
            if os.path.exists(history_file):
                # Find best performance for each model
                model_best = {}
                for record in history_data:
                    model_name = record.get('model_name')
                    accuracy = record.get('accuracy', 0)
                    
                    if model_name not in model_best or accuracy > model_best[model_name]['accuracy']:
                        model_best[model_name] = {
                            'accuracy': accuracy,
                            'f1_score': record.get('f1_score', 0),
                            'iteration': record.get('iteration', 0),
                            'training_time': record.get('training_time', 0)
                        }
                
                # Add explanation at the top
                ws3.merge_cells('A1:F2')
                explanation_cell = ws3.cell(row=1, column=1, 
                    value="🏆 模型对比表说明：\n"
                    "此表格汇总了每个模型在整个训练过程中的最佳表现。\n"
                    "• 模型名称：使用的机器学习算法\n"
                    "• 最佳准确率：模型在整个训练中达到的最高准确率\n"
                    "• 最佳F1分数：模型在整个训练中达到的最高F1分数\n"
                    "• 达到轮次：达到最佳性能的训练轮次\n"
                    "• 训练时间(秒)：达到最佳性能时的训练耗时\n"
                    "• 性能排名：根据最佳准确率从高到低排名\n\n"
                    "💡 决策建议：\n"
                    "1. 排名第1的模型是整体最佳选择\n"
                    "2. 考虑准确率和训练时间的平衡\n"
                    "3. 如果准确率相近，选择训练时间更短的模型")
                explanation_cell.alignment = Alignment(wrap_text=True, vertical='center')
                explanation_cell.font = Font(bold=True, size=11)
                explanation_cell.fill = PatternFill(start_color="FAE6E6", end_color="FAE6E6", fill_type="solid")
                
                # Write headers (starting from row 4)
                headers = ["模型名称", "最佳准确率", "最佳F1分数", "达到轮次", "训练时间(秒)", "性能排名"]
                for col, header in enumerate(headers, 1):
                    cell = ws3.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Sort by accuracy
                sorted_models = sorted(model_best.items(), key=lambda x: x[1]['accuracy'], reverse=True)
                
                row = 5
                for rank, (model_name, metrics) in enumerate(sorted_models, 1):
                    ws3.cell(row=row, column=1, value=model_name)
                    ws3.cell(row=row, column=2, value=metrics['accuracy'])
                    ws3.cell(row=row, column=3, value=metrics['f1_score'])
                    ws3.cell(row=row, column=4, value=metrics['iteration'])
                    ws3.cell(row=row, column=5, value=metrics['training_time'])
                    ws3.cell(row=row, column=6, value=rank)
                    
                    # Add special marking for first place
                    if rank == 1:
                        for col in range(1, 7):
                            cell = ws3.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            cell.font = Font(bold=True)
                    
                    row += 1
                
                # Adjust column widths
                for column in ws3.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    ws3.column_dimensions[column_letter].width = adjusted_width
                
                # Format numbers
                for row in ws3.iter_rows(min_row=5, max_row=row-1, min_col=2, max_col=3):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
            
            # 4. ROC/AUC Summary Sheet with Explanation
            ws4 = wb.create_sheet("ROC_AUC总结")
            
            # Read ROC results
            roc_summary_file = os.path.join("roc_analysis", "ROC_AUC_Summary.json")
            if os.path.exists(roc_summary_file):
                with open(roc_summary_file, 'r', encoding='utf-8') as f:
                    roc_data = json.load(f)
                
                # Add explanation at the top
                ws4.merge_cells('A1:D2')
                explanation_cell = ws4.cell(row=1, column=1, 
                    value="📊 ROC/AUC总结表说明：\n"
                    "此表格汇总了每个模型的ROC曲线分析结果。\n"
                    "• 模型名称：使用的机器学习算法\n"
                    "• Micro平均AUC：考虑所有样本的平均AUC值\n"
                    "• Macro平均AUC：考虑所有类别的平均AUC值\n"
                    "• 类别AUC(平均)：各诊断类别AUC值的平均值\n\n"
                    "📈 AUC值解释：\n"
                    "• 0.9-1.0：极好的区分能力\n"
                    "• 0.8-0.9：良好的区分能力\n"
                    "• 0.7-0.8：中等区分能力\n"
                    "• 0.6-0.7：较差的区分能力\n"
                    "• <0.6：无区分能力")
                explanation_cell.alignment = Alignment(wrap_text=True, vertical='center')
                explanation_cell.font = Font(bold=True, size=11)
                explanation_cell.fill = PatternFill(start_color="E6FAFA", end_color="E6FAFA", fill_type="solid")
                
                # Write headers (starting from row 4)
                headers = ["模型名称", "Micro平均AUC", "Macro平均AUC", "类别AUC(平均)"]
                for col, header in enumerate(headers, 1):
                    cell = ws4.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data (starting from row 5)
                row = 5
                for model_name, result in roc_data.items():
                    if 'error' not in result:
                        ws4.cell(row=row, column=1, value=model_name)
                        ws4.cell(row=row, column=2, value=result.get('micro_auc', 0))
                        ws4.cell(row=row, column=3, value=result.get('macro_auc', 0))
                        
                        # Calculate average class AUC
                        class_aucs = result.get('class_auc', {})
                        if class_aucs:
                            # Extract numeric AUC values (skip 'micro' and 'macro')
                            auc_values = [v for k, v in class_aucs.items() if k not in ['micro', 'macro']]
                            if auc_values:
                                avg_class_auc = np.mean(auc_values)
                                ws4.cell(row=row, column=4, value=avg_class_auc)
                        
                        row += 1
                
                # Adjust column widths
                for column in ws4.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    ws4.column_dimensions[column_letter].width = adjusted_width
                
                # Format numbers
                for row in ws4.iter_rows(min_row=5, max_row=row-1, min_col=2, max_col=4):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.000'
            else:
                print(f"    ⚠️ ROC summary file {roc_summary_file} not found")
            
            # 5. Diagnosis Category Mapping Sheet
            ws5 = wb.create_sheet("诊断类别映射")
            
            # Add explanation at the top
            ws5.merge_cells('A1:C3')
            explanation_cell = ws5.cell(row=1, column=1, 
                value="🔤 诊断类别映射表说明：\n"
                "此表格显示了诊断类别编码与英文名称的对应关系。\n"
                "• 类别编码：机器学习模型中使用的数字编码\n"
                "• 英文名称：诊断类别的英文名称\n"
                "• 中文解释：诊断类别的中文含义\n\n"
                "💡 使用说明：\n"
                "在分析ROC曲线和模型预测结果时，参考此映射表理解各类别的含义。")
            explanation_cell.alignment = Alignment(wrap_text=True, vertical='center')
            explanation_cell.font = Font(bold=True, size=11)
            explanation_cell.fill = PatternFill(start_color="FAF0E6", end_color="FAF0E6", fill_type="solid")
            
            # Write headers (starting from row 5)
            headers = ["类别编码", "英文名称", "中文解释"]
            for col, header in enumerate(headers, 1):
                cell = ws5.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Define diagnosis category mapping with explanations
            diagnosis_mapping = [
                (0, "Hyperopia", "远视"),
                (1, "Emmetropia", "正视"),
                (2, "Pre-myopia", "近视前期"),
                (3, "Mild Myopia", "轻度近视"),
                (4, "Mild Myopia (Long Axial)", "轻度近视（眼轴偏长）"),
                (5, "Moderate Myopia", "中度近视"),
                (6, "High Myopia", "高度近视"),
                (7, "Unclassified", "未分类"),
                (8, "Diagnosis Failed", "诊断失败")
            ]
            
            # Write data (starting from row 6)
            row = 6
            for code, en_name, cn_explanation in diagnosis_mapping:
                ws5.cell(row=row, column=1, value=code)
                ws5.cell(row=row, column=2, value=en_name)
                ws5.cell(row=row, column=3, value=cn_explanation)
                row += 1
            
            # Adjust column widths
            for column in ws5.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws5.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            wb.save(output_file)
            print(f"  ✅ Excel training report generated: {output_file}")
            
            return True
            
        except Exception as e:
            print(f"  ❌ Failed to generate Excel report: {e}")
            traceback.print_exc()
            return False

class ContinuousLearningPipeline:
    """Continuous Learning Pipeline"""
    
    def __init__(self, excel_path: str, max_iterations: int = 10):
        self.excel_path = excel_path
        self.max_iterations = max_iterations
        self.data_processor = EnhancedDataProcessor()
        self.multi_model = MultiModelTraining()
        self.iteration_results = []
        self.full_data = None
        
    # 修改 ContinuousLearningPipeline 中的 run_continuous_training 方法：

def run_continuous_training(self):
    """运行持续训练"""
    print("="*60)
    print("近视诊断模型持续学习管道")
    print(f"最大迭代次数: {self.max_iterations}")
    print("="*60)
    
    try:
        # 1. 加载和预处理数据
        print("\n[阶段1] 数据预处理...")
        df = self.data_processor.load_data(self.excel_path)
        processed_data = self.data_processor.process_data(df)
        
        # 2. 生成患者ID
        processed_data = self.data_processor.generate_patient_ids(processed_data)
        
        # 3. 运行规则诊断
        labeled_data = self.data_processor.run_rule_diagnosis(processed_data)
        self.full_data = labeled_data
        
        # 4. 初始化模型
        print("\n[阶段2] 初始化机器学习模型...")
        self.multi_model.initialize_models()
        
        # 5. 多轮训练
        print("\n[阶段3] 多轮训练...")
        for iteration in range(1, self.max_iterations + 1):
            print(f"\n{'='*50}")
            print(f"第 {iteration}/{self.max_iterations} 轮训练")
            print(f"{'='*50}")
            
            # 使用不同的随机种子确保每次数据分割都不同
            # 使用时间戳确保真正的随机性
            current_time_seed = int(time.time() * 1000) % 10000
            random_seed = current_time_seed + iteration * 100
            
            print(f"  使用随机种子: {random_seed}")
            
            try:
                # 使用不同的随机种子分割数据
                train_data, val_data, test_data = self.data_processor.split_data_by_patient_id(
                    self.full_data,
                    random_state=random_seed  # 传入随机种子
                )
                
                # 检查数据是否有变化
                if iteration > 1:
                    self._check_data_variation(train_data, val_data, test_data, iteration)
                
                # 训练所有模型
                results = self.multi_model.train_and_evaluate_all(
                    train_data, val_data, test_data, iteration
                )
                
                # 对比模型性能
                comparison_df = self.multi_model.compare_models(results)
                
                # 保存本轮结果
                self.iteration_results.append({
                    'iteration': iteration,
                    'results': results,
                    'comparison': comparison_df.to_dict(),
                    'random_seed': random_seed  # 保存随机种子
                })
                
                # 集成预测
                ensemble_pred, ensemble_proba = self.multi_model.ensemble_predict(
                    results, self.multi_model.prepare_features(test_data)
                )
                
                if ensemble_pred is not None:
                    ensemble_accuracy = accuracy_score(test_data['诊断编码'].values, ensemble_pred)
                    print(f"  🤝 集成模型准确率: {ensemble_accuracy:.2%}")
                
                # 保存数据集（第一次迭代）
                if iteration == 1:
                    self.data_processor.save_datasets(train_data, val_data, test_data)
                
                # 等待一下（确保随机种子变化）
                if iteration < self.max_iterations:
                    wait_time = 0.1  # 短暂等待以确保时间戳不同
                    time.sleep(wait_time)
                    
            except Exception as e:
                print(f"  第{iteration}轮训练失败: {e}")
                continue
        
        # 6. 保存所有模型和结果
        print("\n" + "="*60)
        print("[阶段4] 保存模型和训练结果...")
        self.multi_model.save_all_models()
        
        # 保存迭代结果
        self.save_iteration_results()
        
        # 生成最终报告
        self.generate_final_report()
        
        print("\n" + "="*60)
        print("✅ 持续学习训练完成！")
        print("="*60)
        
        return True
        
    except Exception as e:
        print(f"\n❌ 流程执行失败: {e}")
        traceback.print_exc()
        return False

def _check_data_variation(self, train_data, val_data, test_data, iteration):
    """检查数据变化"""
    if iteration > 1:
        # 简单检查：比较样本数量是否有变化
        prev_train_size = getattr(self, '_prev_train_size', 0)
        prev_val_size = getattr(self, '_prev_val_size', 0)
        prev_test_size = getattr(self, '_prev_test_size', 0)
        
        train_change = abs(len(train_data) - prev_train_size)
        val_change = abs(len(val_data) - prev_val_size)
        test_change = abs(len(test_data) - prev_test_size)
        
        if train_change + val_change + test_change > 0:
            print(f"  数据分割变化: 训练集变化{train_change}个样本")
        else:
            print(f"  ⚠️ 数据分割没有变化！")
            print(f"  可能需要检查数据量或随机种子生成")
        
        # 保存当前大小供下次比较
        self._prev_train_size = len(train_data)
        self._prev_val_size = len(val_data)
        self._prev_test_size = len(test_data)
        
    def save_iteration_results(self):
        """Save iteration results"""
        output_file = "iteration_results.json"
        
        # Simplify results for saving
        simplified_results = []
        for iteration_data in self.iteration_results:
            simple_iteration = {
                'iteration': iteration_data['iteration'],
                'model_performance': {}
            }
            
            if 'results' in iteration_data:
                for model_name, result in iteration_data['results'].items():
                    if 'metrics' in result:
                        simple_iteration['model_performance'][model_name] = {
                            'accuracy': result['metrics']['accuracy'],
                            'f1_score': result['metrics']['f1_score'],
                            'training_time': result['metrics']['training_time']
                        }
            
            simplified_results.append(simple_iteration)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(simplified_results, f, ensure_ascii=False, indent=2)
        
        print(f"  Iteration results saved to: {output_file}")
    
    def generate_final_report(self):
        """Generate final report"""
        print("\n" + "="*60)
        print("Final Training Report")
        print("="*60)
        
        # Get best models
        best_models = self.multi_model.performance_tracker.get_best_models()
        
        print("\n🏆 Best performance for each model:")
        for model_name, best_info in best_models.items():
            metrics = best_info['metrics']
            print(f"\n  {model_name}:")
            print(f"    Iteration: {best_info['iteration']}")
            print(f"    Accuracy: {metrics['accuracy']:.2%}")
            print(f"    F1 Score: {metrics['f1_score']:.2%}")
            print(f"    Validation Accuracy: {metrics.get('val_accuracy', 0):.2%}")
            print(f"    Training Time: {metrics['training_time']:.2f}s")
            print(f"    Time: {best_info['timestamp']}")
        
        # Summary
        print("\n📊 Training Summary:")
        print(f"  Total iterations: {len(self.iteration_results)}")
        print(f"  Number of trained models: {len(self.multi_model.models)}")
        print(f"  Total data volume: {len(self.full_data) if self.full_data is not None else 0}")
        
        # Generate recommendations
        print("\n💡 Recommendations:")
        if best_models:
            best_model_name = max(best_models.keys(), 
                                key=lambda x: best_models[x]['metrics']['accuracy'])
            print(f"  Recommended model: {best_model_name}")
            print(f"  Expected accuracy: {best_models[best_model_name]['metrics']['accuracy']:.2%}")
        
        print("\n📁 Generated files:")
        print("  1. saved_models/ - All trained models")
        print("  2. model_history.json - Model training history")
        print("  3. iteration_results.json - Iteration results")
        print("  4. training_set.xlsx, validation_set.xlsx, test_set.xlsx - Dataset splits")

class EnhancedContinuousLearningPipeline(ContinuousLearningPipeline):
    """Enhanced Continuous Learning Pipeline (with testing and reporting)"""
    
    def __init__(self, excel_path: str, max_iterations: int = 10):
        super().__init__(excel_path, max_iterations)
        self.roc_results = {}
        self.final_test_results = {}
        
    def run_continuous_training(self):
        """Run enhanced training pipeline"""
        # Call parent class training method
        success = super().run_continuous_training()
        
        if success:
            # Add extra functionality
            self.generate_excel_reports()
            self.test_real_cases()
            
        return success
    
    def generate_excel_reports(self):
        """Generate Excel reports"""
        print("\n" + "="*60)
        print("[Phase 5] Generating Excel reports and ROC analysis")
        print("="*60)
        
        # Generate Excel report
        ExcelReportGenerator.generate_training_report(
            os.path.join("saved_models", "model_history.json"),
            "iteration_results.json",
            "训练报告.xlsx"
        )
        
        # Analyze ROC curves (using last iteration results)
        if hasattr(self.multi_model, 'analyze_roc_for_all_models') and self.iteration_results:
            last_iteration = self.iteration_results[-1]
            if 'results' in last_iteration:
                self.roc_results = self.multi_model.analyze_roc_for_all_models(
                    last_iteration['results'],
                    label_encoder=self.data_processor.label_encoder,
                    output_dir="roc_analysis"
                )
    
    def test_real_cases(self):
        """Test real cases"""
        print("\n" + "="*60)
        print("[Phase 6] Real Case Testing")
        print("="*60)
        
        # Select some cases from test set for testing
        if hasattr(self, 'full_data') and self.full_data is not None:
            # Find patients in test set
            test_patients = self.full_data[self.full_data['patient_id'].str.contains('TEST', na=False)]
            if len(test_patients) == 0:
                # If no TEST label, randomly select
                test_patients = self.full_data.sample(min(5, len(self.full_data)))
            
            print(f"  Selected {len(test_patients)} cases for testing:")
            
            for idx, patient in test_patients.iterrows():
                self.test_single_case(patient)
                print("-" * 50)
        
        # Interactive testing
        self.interactive_testing()
    
    def test_single_case(self, patient_data):
        """Test single case"""
        try:
            print(f"\n  Testing case {patient_data['姓名']} ({patient_data['眼别']}):")
            print(f"    Age: {patient_data['年龄']} years")
            print(f"    Gender: {patient_data['性别']}")
            print(f"    Vision: {patient_data['视力']}")
            print(f"    Refraction: {patient_data['原始_验光']}D")
            print(f"    Axial length: {patient_data['眼轴']}mm")
            print(f"    Axial ratio: {patient_data['轴率比']}")
            print(f"    Corneal curvature: {patient_data['角膜曲率']}D")
            
            # Use rule-based system for diagnosis
            rule_diagnosis = self.data_processor.rule_classifier.diagnose({
                'age': patient_data['年龄'],
                'se': patient_data['验光'],
                'al': patient_data['眼轴'],
                'corneal_curv': patient_data['角膜曲率'],
                'va': patient_data['视力'],
                'is_cycloplegic': False
            })
            
            print(f"    Rule-based diagnosis: {rule_diagnosis['stage']}")
            
            # Use machine learning models for diagnosis
            ml_predictions = {}
            
            # Prepare features
            patient_features = self.multi_model.prepare_features(pd.DataFrame([patient_data]))
            
            # Predict using each model
            for model_name, model_wrapper in self.multi_model.models.items():
                try:
                    model = model_wrapper.base_model
                    prediction = model.predict(patient_features)[0]
                    
                    # If probabilities available, also display
                    if hasattr(model, 'predict_proba'):
                        proba = model.predict_proba(patient_features)[0]
                        max_prob = np.max(proba)
                        ml_predictions[model_name] = {
                            'prediction': prediction,
                            'probability': max_prob,
                            'all_probs': proba.tolist()
                        }
                    else:
                        ml_predictions[model_name] = {
                            'prediction': prediction,
                            'probability': None
                        }
                    
                except Exception as e:
                    print(f"    ❌ {model_name} prediction failed: {e}")
            
            # Display machine learning prediction results
            print(f"    Machine learning diagnosis:")
            
            for model_name, pred_info in ml_predictions.items():
                if 'prediction' in pred_info:
                    # Decode prediction result
                    diagnosis_label = "Unknown"
                    try:
                        if pred_info['prediction'] in self.data_processor.label_encoder.classes_:
                            diagnosis_label = self.data_processor.label_encoder.inverse_transform(
                                [pred_info['prediction']]
                            )[0]
                    except:
                        pass
                    
                    prob_text = f" (confidence: {pred_info['probability']:.2%})" if pred_info['probability'] else ""
                    print(f"      {model_name}: {diagnosis_label}{prob_text}")
            
            # True diagnosis (if available)
            if '诊断结果' in patient_data:
                print(f"    True diagnosis: {patient_data['诊断结果']}")
                
                # Check if matches
                rule_match = rule_diagnosis['stage'] == patient_data['诊断结果']
                print(f"    Rule-based diagnosis match: {'✅' if rule_match else '❌'}")
                
                # Check machine learning diagnosis match
                for model_name, pred_info in ml_predictions.items():
                    if 'prediction' in pred_info:
                        ml_diagnosis = "Unknown"
                        try:
                            if pred_info['prediction'] in self.data_processor.label_encoder.classes_:
                                ml_diagnosis = self.data_processor.label_encoder.inverse_transform(
                                    [pred_info['prediction']]
                                )[0]
                        except:
                            pass
                        
                        ml_match = ml_diagnosis == patient_data['诊断结果']
                        print(f"    {model_name} match: {'✅' if ml_match else '❌'}")
            
            return {
                'patient_info': {
                    '姓名': patient_data['姓名'],
                    '年龄': patient_data['年龄'],
                    '眼别': patient_data['眼别']
                },
                'rule_diagnosis': rule_diagnosis['stage'],
                'ml_predictions': ml_predictions,
                'true_diagnosis': patient_data.get('诊断结果', 'Unknown')
            }
            
        except Exception as e:
            print(f"    Case testing failed: {e}")
            return None
    
    def interactive_testing(self):
        """Interactive testing"""
        print("\n  Interactive testing:")
        print("  1. Use random cases from test data")
        print("  2. Manually enter case information")
        print("  3. Batch testing")
        print("  4. Exit testing")
        
        try:
            choice = input("  Please choose testing method (1-4): ")
            
            if choice == '1':
                self.test_random_cases()
            elif choice == '2':
                self.test_manual_input()
            elif choice == '3':
                self.test_batch_cases()
            elif choice == '4':
                print("  Exiting testing")
            else:
                print("  Invalid choice")
                
        except Exception as e:
            print(f"  Interactive testing failed: {e}")
    
    def test_random_cases(self, n_cases=3):
        """Randomly test cases"""
        if hasattr(self, 'full_data') and self.full_data is not None:
            random_cases = self.full_data.sample(min(n_cases, len(self.full_data)))
            
            print(f"\n  Randomly testing {len(random_cases)} cases:")
            
            for idx, case in random_cases.iterrows():
                result = self.test_single_case(case)
                self.final_test_results[f"random_{idx}"] = result
                print("-" * 50)
    
    def test_manual_input(self):
        """Manually input case information for testing"""
        print("\n  Manual case information input:")
        
        try:
            name = input("  Name: ") or "Test Patient"
            age = int(input("  Age: ") or "8")
            gender = input("  Gender (男/女): ") or "男"
            eye = input("  Eye (左眼/右眼): ") or "右眼"
            vision = float(input("  Vision: ") or "0.8")
            refraction = float(input("  Refraction (D): ") or "-1.5")
            axial_length = float(input("  Axial length (mm): ") or "24.5")
            axial_ratio = float(input("  Axial ratio: ") or "3.2")
            corneal_curv = float(input("  Corneal curvature (D): ") or "43.2")
            
            # Create simulated patient data
            patient_data = {
                '姓名': name,
                '年龄': age,
                '性别': gender,
                '眼别': eye,
                '视力': vision,
                '验光': refraction,
                '眼轴': axial_length,
                '轴率比': axial_ratio,
                '角膜曲率': corneal_curv,
                '原始_年龄': str(age),
                '原始_验光': str(refraction),
                'patient_key': f"{name}_{1 if gender == '女' else 0}_{age}",
                'patient_id': PatientIDGenerator.generate_patient_id({'姓名': name, '性别': gender, '年龄': age})
            }
            
            result = self.test_single_case(patient_data)
            self.final_test_results["manual_input"] = result
            
        except Exception as e:
            print(f"  Manual input failed: {e}")
    
    def test_batch_cases(self):
        """Batch testing (read from file)"""
        print("\n  Batch testing functionality will be implemented in future versions")
        print("  You can create CSV or Excel files containing multiple case information")

# Main program
if __name__ == "__main__":
    EXCEL_FILE = "博士数据收集适用.xlsx"  # Replace with your Excel path
    
    print("Myopia Diagnosis Continuous Learning System - Enhanced Version")
    print("="*60)
    print("Supported models: RandomForest, LogisticRegression, SVM, GradientBoosting" + (", XGBoost" if XGBOOST_AVAILABLE else ""))
    print("="*60)
    
    # Get number of iterations
    try:
        max_iterations = int(input("Please enter number of training iterations (default 5): ") or "5")
    except:
        max_iterations = 5
    
    # Run enhanced continuous learning pipeline
    pipeline = EnhancedContinuousLearningPipeline(EXCEL_FILE, max_iterations)
    success = pipeline.run_continuous_training()
    
    if success:
        print("\n" + "="*60)
        print("🎉 Enhanced training completed! All functions executed.")
        print("="*60)
        
        print("\n📁 Generated files and reports:")
        print("  1. saved_models/ - All trained models")
        print("  2. 训练报告.xlsx - Excel training report (with history, iterations, comparison)")
        print("  3. roc_analysis/ - ROC curves and AUC analysis results")
        print("  4. iteration_results.json - Iteration results")
        print("  5. 训练集.xlsx, 验证集.xlsx, 测试集.xlsx - Dataset splits")
        print("  6. model_history.json - Model training history")
        
        # Display best model
        if hasattr(pipeline.multi_model, 'performance_tracker'):
            best_models = pipeline.multi_model.performance_tracker.get_best_models()
            if best_models:
                print("\n🏆 Best model summary:")
                best_model_name = max(best_models.keys(), 
                                    key=lambda x: best_models[x]['metrics']['accuracy'])
                print(f"  Recommended model: {best_model_name}")
                print(f"  Best accuracy: {best_models[best_model_name]['metrics']['accuracy']:.2%}")
                
        # ROC results summary
        if pipeline.roc_results:
            print("\n📊 ROC/AUC Summary:")
            for model_name, result in pipeline.roc_results.items():
                if 'error' not in result:
                    print(f"  {model_name}:")
                    print(f"    Macro-average AUC: {result.get('macro_auc', 0):.3f}")
                    print(f"    Micro-average AUC: {result.get('micro_auc', 0):.3f}")
    else:
        print("\n❌ Issues occurred during training, please check logs.")
